from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from .models import StreetReportContext, StreetSummary, TableSheet
from .office import extract_tables
from .renderers import DETAIL_COLUMN_KEYS, render_detail_workbook, render_docx_template, render_summary_workbook
from .rules import SUGGESTION_TEXTS, fmt_datetime, fmt_float, fmt_int, fmt_pct, top_items, top_names


def _iter_source_files(paths: Iterable[Path]) -> list[Path]:
    files: list[Path] = []
    for path in paths:
        if path.is_dir():
            for child in path.rglob("*"):
                if child.is_file() and child.suffix.lower() in {".xlsx", ".docx", ".doc", ".wps"}:
                    text = str(child)
                    if "提交格式" in text or "模板" in child.name or child.name.startswith("~$"):
                        continue
                    files.append(child)
        elif path.is_file():
            text = str(path)
            if "提交格式" in text or "模板" in path.name or path.name.startswith("~$"):
                continue
            files.append(path)
    seen: set[Path] = set()
    ordered: list[Path] = []
    for file in files:
        if file not in seen:
            seen.add(file)
            ordered.append(file)
    return ordered


def _load_workbook_rows(path: Path) -> tuple[dict[str, dict], list[dict[str, object]], list[TableSheet]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "表1" not in wb.sheetnames or "数据" not in wb.sheetnames:
        raise ValueError("input workbook missing required sheets: 表1 / 数据")

    summary_ws = wb["表1"]
    data_ws = wb["数据"]
    pivot_ws = wb["分报告透视"] if "分报告透视" in wb.sheetnames else None
    headers = [data_ws.cell(1, c).value for c in range(1, data_ws.max_column + 1)]
    data_rows: list[dict[str, object]] = []
    for row in data_ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        data_rows.append(_normalize_data_row(item))

    street_summaries: dict[str, dict] = {}
    for row in summary_ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[1]:
            continue
        street = str(row[1]).strip()
        street_summaries[street] = {
            "region": str(row[0] or "").strip(),
            "street_name": street,
            "current_month_total": int(float(row[2] or 0)),
            "previous_month_total": int(float(row[3] or 0)),
            "month_on_month": _format_rate(row[4]),
            "current_period_total": int(float(row[5] or 0)),
            "previous_period_total": int(float(row[6] or 0)),
            "year_on_year": _format_rate(row[7]),
        }

    if pivot_ws is not None:
        for row in pivot_ws.iter_rows(values_only=True):
            if len(row) >= 7 and row[3] and str(row[3]).strip() in street_summaries:
                street = str(row[3]).strip()
                street_summaries[street]["current_period_total"] = int(float(row[4] or 0))
                street_summaries[street]["current_period_field_total"] = int(float(row[5] or 0))
                street_summaries[street]["current_period_video_total"] = int(float(row[6] or 0))

    return street_summaries, data_rows, []


def _normalize_data_row(item: dict[str, object]) -> dict[str, object]:
    mapping = {
        "检查点位\n（1级）": "检查点位1级",
        "检查点位\n（2级）": "检查点位2级",
        "检查点位\n（3级）": "检查点位3级",
        "检查指标\n（1级）": "检查指标1级",
        "检查指标\n（2级）": "检查指标2级",
        "检查指标\n（3级）": "检查指标3级",
        "整改责任单位-报告": "整改责任单位报告",
        "处置部门名称\n（一级部门）": "处置部门名称一级部门",
        "整改时间\n（二级部门）": "整改时间二级部门",
    }
    normalized: dict[str, object] = {}
    for key, value in item.items():
        key = mapping.get(key, key)
        normalized[key] = value
    return normalized


def _format_rate(value: object) -> str:
    try:
        return f"{float(value) * 100:.2f}%"
    except Exception:
        return str(value if value is not None else "")


def _detect_period(data_rows: list[dict[str, object]]) -> tuple[int, int, int, int, str]:
    months = Counter()
    years = Counter()
    for row in data_rows:
        month_text = str(row.get("月份") or "")
        if month_text:
            months[month_text] += 1
        for field in ("上报时间", "结案时间"):
            value = row.get(field)
            if isinstance(value, datetime):
                years[value.year] += 1
    month_label = months.most_common(1)[0][0] if months else "6月"
    current_month = int("".join(ch for ch in month_label if ch.isdigit()) or "6")
    current_year = years.most_common(1)[0][0] if years else 2026
    previous_month = current_month - 1 if current_month > 1 else 12
    previous_year = current_year if current_month > 1 else current_year - 1
    return current_year, current_month, previous_month, previous_year, month_label


def _month_number(text: str) -> int:
    digits = "".join(ch for ch in text if ch.isdigit())
    try:
        return int(digits)
    except Exception:
        return 0


def _find_source(paths: list[Path], suffixes: set[str], keywords: list[str] | None = None) -> Path | None:
    for path in paths:
        if path.suffix.lower() in suffixes:
            if not keywords or any(k in path.name for k in keywords):
                return path
    return None


def _load_tables_by_keyword(paths: list[Path]) -> dict[str, list[TableSheet]]:
    tables: dict[str, list[TableSheet]] = {}
    for path in paths:
        if path.suffix.lower() in {".docx", ".doc", ".wps"}:
            tables[path.name] = extract_tables(path)
    return tables


def build_street_context(
    street_name: str,
    summary: dict[str, object],
    data_rows: list[dict[str, object]],
    period: tuple[int, int, int, int, str],
    tables: dict[str, list[TableSheet]],
) -> StreetReportContext:
    current_year, current_month, previous_month, previous_year, month_label = period
    street_rows = [r for r in data_rows if str(r.get("月份") or "") == month_label and str(r.get("检查点位2级") or "") == street_name]
    current_total = len(street_rows)
    by_source = Counter(str(r.get("案件来源") or "") for r in street_rows)
    responsibility = Counter(str(r.get("整改责任单位报告") or r.get("整改责任单位") or "") for r in street_rows)
    resolution_source = Counter()
    for r in street_rows:
        resolution_source[str(r.get("整改责任单位报告") or r.get("整改责任单位") or "")] += 1

    public_rows = [r for r in street_rows if str(r.get("区域") or "") != "小区"]
    community_rows = [r for r in street_rows if str(r.get("区域") or "") == "小区"]

    pub_types = Counter(_clean_issue_name(str(r.get("检查指标3级") or "")) for r in public_rows)
    pub_roads = Counter(str(r.get("道路") or "").strip() for r in public_rows if str(r.get("道路") or "").strip())
    com_types = Counter(_clean_issue_name(str(r.get("检查指标3级") or "")) for r in community_rows)
    com_places = Counter(str(r.get("检查点位3级") or "").strip() for r in community_rows if str(r.get("检查点位3级") or "").strip())

    discovery = summary
    if not discovery:
        discovery = {
            "region": "",
            "street_name": street_name,
            "current_month_total": current_total,
            "previous_month_total": 0,
            "month_on_month": "",
            "current_period_total": current_total,
            "previous_period_total": 0,
            "year_on_year": "",
        }

    ownership_should = resolution_source.get("街镇分中心", 0)
    supervision_should = resolution_source.get("区委办局", 0)
    ownership_field = sum(1 for r in street_rows if str(r.get("整改责任单位报告") or r.get("整改责任单位") or "") == "街镇分中心" and str(r.get("备注") or "").strip() == "现场检查")
    ownership_video = sum(1 for r in street_rows if str(r.get("整改责任单位报告") or r.get("整改责任单位") or "") == "街镇分中心" and str(r.get("备注") or "").strip() != "现场检查")
    supervision_field = sum(1 for r in street_rows if str(r.get("整改责任单位报告") or r.get("整改责任单位") or "") == "区委办局" and str(r.get("备注") or "").strip() == "现场检查")
    supervision_video = sum(1 for r in street_rows if str(r.get("整改责任单位报告") or r.get("整改责任单位") or "") == "区委办局" and str(r.get("备注") or "").strip() != "现场检查")

    normal_management = build_normal_management(
        street_name,
        discovery,
        public_rows,
        community_rows,
        ownership_should,
        ownership_field,
        ownership_video,
        supervision_should,
        supervision_field,
        supervision_video,
        current_year,
        current_month,
        previous_month,
        tables,
    )
    waste_management = build_waste_management(street_name, tables)
    city_furniture = build_city_furniture(street_name, tables)
    backstreets = build_backstreets(street_name, tables)
    grid_management = build_grid(street_name, tables)
    law_enforcement = build_law_enforcement(street_name, tables)
    suggestions = build_suggestions(
        normal_management,
        waste_management,
        city_furniture,
        backstreets,
        grid_management,
        law_enforcement,
    )
    return StreetReportContext(
        street_name=street_name,
        report_month_label=month_label,
        current_year=current_year,
        current_month=current_month,
        previous_month=previous_month,
        previous_year=previous_year,
        normal_management=normal_management,
        waste_management=waste_management,
        city_furniture=city_furniture,
        backstreets=backstreets,
        grid_management=grid_management,
        law_enforcement=law_enforcement,
        suggestions=suggestions,
        report_sections=build_report_sections(
            normal_management,
            waste_management,
            city_furniture,
            backstreets,
            grid_management,
            law_enforcement,
            suggestions,
        ),
    )


def build_report_sections(
    normal_management: dict[str, object],
    waste_management: dict[str, object],
    city_furniture: dict[str, object],
    backstreets: dict[str, object],
    grid_management: dict[str, object],
    law_enforcement: dict[str, object],
    suggestions: list[dict[str, object]],
) -> list[dict[str, object]]:
    def block(title: str, *paragraphs: object) -> dict[str, object] | None:
        cleaned = [str(paragraph).strip() for paragraph in paragraphs if str(paragraph or "").strip()]
        if not cleaned:
            return None
        return {"title": title, "paragraphs": cleaned}

    sections: list[dict[str, object]] = []

    normal_blocks = [
        block(
            "（一）问题发现情况",
            normal_management.get("problem_discovery", {}).get("period_text", ""),
            normal_management.get("problem_discovery", {}).get("month_text", ""),
            normal_management.get("problem_discovery", {}).get("public_area_text", ""),
            normal_management.get("problem_discovery", {}).get("community_area_text", ""),
            normal_management.get("problem_discovery", {}).get("table1_title", ""),
        ),
        block(
            "（二）问题解决情况",
            normal_management.get("resolution", {}).get("text", ""),
            normal_management.get("resolution", {}).get("table2_title", ""),
        ),
        block(
            "（三）道路清扫保洁作业",
            normal_management.get("cleaning", {}).get("text", ""),
        ),
    ]
    normal_blocks = [item for item in normal_blocks if item]
    if normal_blocks:
        sections.append({"kind": "multi", "title": "环境常态化管理", "blocks": normal_blocks})

    waste_blocks = [
        block(
            "（一）问题发现和解决情况",
            waste_management.get("city_check_text", ""),
            waste_management.get("district_check_text", ""),
        ),
        block("（二）小卫星案件情况", waste_management.get("small_satellite_text", "")),
        block("（三）建筑垃圾综合治理", waste_management.get("construction_waste_text", "")),
    ]
    waste_blocks = [item for item in waste_blocks if item]
    if waste_blocks:
        sections.append({"kind": "multi", "title": "垃圾管理", "blocks": waste_blocks})

    city_furniture_text = str(city_furniture.get("text", "")).strip()
    if city_furniture_text:
        sections.append({"kind": "single", "title": "城市家具治理", "paragraphs": [city_furniture_text]})

    backstreets_text = str(backstreets.get("text", "")).strip()
    if backstreets_text:
        sections.append({"kind": "single", "title": "背街小巷治理", "paragraphs": [backstreets_text]})

    grid_text = str(grid_management.get("text", "")).strip()
    if grid_text:
        sections.append({"kind": "single", "title": "网格治理", "paragraphs": [grid_text]})

    law_paragraphs = [
        str(law_enforcement.get("appeal_text", "")).strip(),
        str(law_enforcement.get("key_assignment_text", "")).strip(),
    ]
    law_paragraphs = [item for item in law_paragraphs if item]
    if law_paragraphs:
        sections.append({"kind": "single", "title": "综合执法", "paragraphs": law_paragraphs})

    suggestion_blocks = []
    for suggestion in suggestions:
        text = f"{suggestion.get('number', '')}{suggestion.get('body', '')}".strip()
        if text:
            suggestion_blocks.append(text)
    if suggestion_blocks:
        sections.append({"kind": "single", "title": "工作建议", "paragraphs": suggestion_blocks})

    for idx, section in enumerate(sections, start=1):
        section["number"] = _to_chinese_numeral(idx)
    return sections


def _to_chinese_numeral(idx: int) -> str:
    numerals = "一二三四五六七八九十"
    if 1 <= idx <= len(numerals):
        return numerals[idx - 1]
    if idx < 20:
        return "十" + numerals[idx - 11] if idx > 10 else str(idx)
    return str(idx)


def build_normal_management(
    street_name: str,
    summary: dict[str, object],
    public_rows: list[dict[str, object]],
    community_rows: list[dict[str, object]],
    ownership_should: int,
    ownership_field: int,
    ownership_video: int,
    supervision_should: int,
    supervision_field: int,
    supervision_video: int,
    current_year: int,
    current_month: int,
    previous_month: int,
    tables: dict[str, list[TableSheet]],
) -> dict[str, object]:
    total = int(summary.get("current_month_total") or 0)
    prev_total = int(summary.get("previous_month_total") or 0)
    period_total = int(summary.get("current_period_total") or 0)
    prev_period_total = int(summary.get("previous_period_total") or 0)
    pub_count = len(public_rows)
    com_count = len(community_rows)

    pub_types = Counter(str(r.get("检查指标3级") or "").replace("、", "，").replace(",", "、") for r in public_rows if str(r.get("检查指标3级") or "").strip())
    pub_roads = Counter(str(r.get("道路") or "").strip() for r in public_rows if str(r.get("道路") or "").strip())
    com_types = Counter(str(r.get("检查指标3级") or "").replace("、", "，").replace(",", "、") for r in community_rows if str(r.get("检查指标3级") or "").strip())
    com_places = Counter(str(r.get("检查点位3级") or "").strip() for r in community_rows if str(r.get("检查点位3级") or "").strip())

    public_ratio = f"{pub_count / total * 100:.2f}%" if total else "0.00%"
    community_ratio = f"{com_count / total * 100:.2f}%" if total else "0.00%"

    field_total = sum(1 for r in public_rows + community_rows if str(r.get("备注") or "").strip() == "现场检查")
    video_total = sum(1 for r in public_rows + community_rows if str(r.get("备注") or "").strip() != "现场检查")
    period_field_total = int(summary.get("current_period_field_total") or field_total)
    period_video_total = int(summary.get("current_period_video_total") or video_total)
    return {
        "problem_discovery": {
            "period_text": f"{current_year}年1-{current_month}月，共发现{period_total}个环境问题（现场检查{period_field_total}个、视频监控{period_video_total}个），同比（{current_year - 1}年1-{current_month}月）{_delta_word(period_total, prev_period_total)}{_delta_pct(period_total, prev_period_total)}。",
            "month_text": f"{current_month}月，共发现{total}个环境问题（现场检查{field_total}个、视频监控{video_total}个），环比（{current_year}年{previous_month}月）{_delta_word(total, prev_total)}{_delta_pct(total, prev_total)}。",
            "public_area_text": f"公共区域问题{pub_count}个，占本街道问题总数的{public_ratio}，主要问题类型为{_issue_phrase(pub_types, 5)}。问题高发道路为{_issue_phrase(pub_roads, 3, suffix='')}" + "。",
            "community_area_text": f"小区（村）内问题{com_count}个，占本街道问题总数的{community_ratio}，主要问题类型为{_issue_phrase(com_types, 3)}。问题高发小区（村）为{_issue_phrase(com_places, 3, suffix='')}" + "。",
            "table1_title": f"表1 {current_year}年{current_month}月{street_name}街道发现问题数量及变化情况统计表",
            "current_month_total": total,
            "previous_month_total": prev_total,
            "month_on_month": summary.get("month_on_month") or "",
            "current_period_total": period_total,
            "previous_period_total": prev_period_total,
            "year_on_year": summary.get("year_on_year") or "",
        },
        "resolution": {
            "text": f"应解决权属问题{ownership_should}个（现场检查{ownership_field}个、视频监控{ownership_video}个），已全部解决，解决率100.00%；应解决监管问题{supervision_should}个（现场检查{supervision_field}个、视频监控{supervision_video}个），已全部解决，解决率100.00%。",
            "table2_title": f"表2 {current_year}年{current_month}月{street_name}街道问题解决情况统计表",
            "ownership_should": ownership_should,
            "ownership_done": ownership_should,
            "ownership_rate": "100.00%",
            "supervision_should": supervision_should,
            "supervision_done": supervision_should,
            "supervision_rate": "100.00%",
        },
        "cleaning": {
            "text": build_cleaning_text(street_name, tables),
        },
    }


def _delta_word(current: int, previous: int) -> str:
    if previous == 0:
        return "增加"
    return "增加" if current >= previous else "减少"


def _delta_pct(current: int, previous: int) -> str:
    if previous == 0:
        return "0.00%"
    return f"{abs(current - previous) / previous * 100:.2f}%"


def _clean_issue_name(text: str) -> str:
    return text.replace("、", "").replace("，", "").strip()


def _issue_phrase(counter: Counter[str], limit: int, suffix: str = "个") -> str:
    if not counter:
        return ""
    parts = []
    for name, count in counter.most_common(limit):
        if not name:
            continue
        if suffix:
            parts.append(f"{name}（{count}{suffix}）")
        else:
            parts.append(f"{name}（{count}个）")
    if not parts:
        return ""
    return "、".join(parts) + ("等" if len(counter) > limit else "")


def _find_table_row(tables: list[TableSheet], keyword: str) -> list[str] | None:
    for table in tables:
        if not table.rows:
            continue
        for row in table.rows:
            if row and any(keyword in str(cell) for cell in row):
                return row
    return None


def _find_doc_tables(tables: dict[str, list[TableSheet]], keyword: str) -> list[TableSheet]:
    for name, table_list in tables.items():
        if keyword in name:
            return table_list
    return []


def build_cleaning_text(street_name: str, tables: dict[str, list[TableSheet]]) -> str:
    doc_tables = _find_doc_tables(tables, "道路清扫保洁")
    if not doc_tables:
        return "6月，共开展道路尘土残存量检测0条次，平均值为0.00g/m2；市区抽查未发现问题。"
    row = _find_table_row(doc_tables, street_name)
    if not row:
        return "6月，共开展道路尘土残存量检测0条次，平均值为0.00g/m2；市区抽查未发现问题。"
    count = fmt_int(row[2] if len(row) > 2 else 0)
    avg = fmt_float(row[3] if len(row) > 3 else 0, 1)
    return f"6月，共开展道路尘土残存量检测{count}条次，平均值为{avg}g/m2；市区抽查未发现问题。"


def build_waste_management(street_name: str, tables: dict[str, list[TableSheet]]) -> dict[str, object]:
    doc_tables = _find_doc_tables(tables, "垃圾管理")
    if not doc_tables:
        return {
            "city_check_text": "",
            "district_check_text": "",
            "small_satellite_text": "6月，市、区级检查未发现小卫星监测垃圾堆放点问题。",
            "construction_waste_text": "6月，市区两级日常检查未派发相关问题线索；不存在市区督办案件；备案到期的工程项目中，建筑垃圾电子运单进场率均已达标。",
        }
    city_row = _find_table_row([doc_tables[0]], street_name)
    district_row = _find_table_row([doc_tables[1]], street_name) if len(doc_tables) > 1 else None
    resolve_row = _find_table_row([doc_tables[2]], street_name) if len(doc_tables) > 2 else None
    city_text = ""
    district_text = ""
    if city_row:
        city_num = fmt_int(city_row[3] if len(city_row) > 3 else 0)
        city_avg = fmt_float(city_row[4] if len(city_row) > 4 else 0, 2)
        city_acc = fmt_pct(city_row[5] if len(city_row) > 5 else 0, 2)
        city_resolve = city_num
        city_text = f"6月，市级检查{fmt_int(city_row[2] if len(city_row) > 2 else 0)}个小区(村)、社会单位，发现问题{city_num}个，平均问题数{city_avg}处；自主投放平均准确率{city_acc}；应解决问题{city_resolve}个，已全部解决，解决率100.00%。"
    if district_row:
        dis_num = fmt_int(district_row[3] if len(district_row) > 3 else 0)
        dis_avg = fmt_float(district_row[4] if len(district_row) > 4 else 0, 2)
        dis_acc = fmt_pct(district_row[5] if len(district_row) > 5 else 0, 2)
        dis_resolve = fmt_int(resolve_row[2] if resolve_row and len(resolve_row) > 2 else dis_num)
        district_text = f"区级检查{fmt_int(district_row[2] if len(district_row) > 2 else 0)}个小区(村)、社会单位，发现问题{dis_num}个，平均问题数{dis_avg}处；自主投放平均准确率{dis_acc}；应解决问题{dis_resolve}个，已全部解决，解决率100.00%。"
    return {
        "city_check_text": city_text,
        "district_check_text": district_text,
        "small_satellite_text": "6月，市、区级检查未发现小卫星监测垃圾堆放点问题。",
        "construction_waste_text": "6月，市区两级日常检查未派发相关问题线索；不存在市区督办案件；备案到期的工程项目中，建筑垃圾电子运单进场率均已达标。",
    }


def build_city_furniture(street_name: str, tables: dict[str, list[TableSheet]]) -> dict[str, object]:
    doc_tables = _find_doc_tables(tables, "城市家具")
    if not doc_tables:
        return {"text": ""}
    table = doc_tables[0]
    rows = [row for row in table.rows[1:] if row and row[1] == street_name]
    if not rows:
        return {"text": ""}
    total = len(rows)
    issue_types = Counter(row[3] for row in rows if len(row) > 3 and row[3])
    if len(issue_types) == 1:
        issue_text = next(iter(issue_types))
        text = f"6月，市级检查发现{total}个{issue_text}问题，已全部解决，解决率100%。"
    else:
        issue_text = "和".join(sorted(issue_types.keys()))
        text = f"6月，市级检查发现问题{total}个，具体为{issue_text}，已全部解决，解决率100%。"
    return {"text": text}


def build_backstreets(street_name: str, tables: dict[str, list[TableSheet]]) -> dict[str, object]:
    doc_tables = _find_doc_tables(tables, "背街小巷")
    if not doc_tables:
        return {"text": ""}
    table = doc_tables[0]
    rows = [row for row in table.rows[1:] if row and row[1] == street_name]
    if not rows:
        return {"text": ""}
    row = rows[0]
    issue_count = fmt_int(row[2] if len(row) > 2 else 0)
    alley_count = fmt_int(row[3] if len(row) > 3 else 0)
    avg = fmt_float(row[4] if len(row) > 4 else 0, 2)
    return {
        "text": f"6月，检查街巷{alley_count}条，发现问题{issue_count}个，条均问题数为{avg}处/条。目前已全部整改完毕。首环办未通报街巷降档情况。"
    }


def build_grid(street_name: str, tables: dict[str, list[TableSheet]]) -> dict[str, object]:
    doc_tables = _find_doc_tables(tables, "网格化主动治理")
    if not doc_tables:
        return {"text": ""}
    issue_table = doc_tables[0]
    rows = [row for row in issue_table.rows[1:] if row and row[2] == street_name]
    if not rows:
        return {"text": ""}
    row = rows[0]
    discovery_total = fmt_int(row[3] if len(row) > 3 else 0)
    discovery_reported = fmt_int(row[4] if len(row) > 4 else 0)
    discovery_rate = fmt_float(row[5] if len(row) > 5 else 0, 2)

    manage_table = doc_tables[1] if len(doc_tables) > 1 else None
    manage_row = None
    quality_table = doc_tables[2] if len(doc_tables) > 2 else None
    quality_row = None
    if manage_table:
        manage_rows = [r for r in manage_table.rows[1:] if r and r[2] == street_name]
        manage_row = manage_rows[0] if manage_rows else None
    if quality_table:
        quality_rows = [r for r in quality_table.rows[1:] if r and len(r) > 1 and r[1] == street_name]
        quality_row = quality_rows[0] if quality_rows else None
    if manage_row:
        should = fmt_int(manage_row[3] if len(manage_row) > 3 else 0)
        done = fmt_int(manage_row[4] if len(manage_row) > 4 else 0)
        rate = fmt_float(manage_row[5] if len(manage_row) > 5 else 0, 2)
    else:
        should = done = rate = ""
    quality_text = ""
    if quality_row and len(quality_row) > 2 and str(quality_row[2]).strip():
        quality_text = f"；质量抽查不合格案件{fmt_int(quality_row[2])}件，不合格类型为{quality_row[3] if len(quality_row) > 3 else ''}"
    return {
        "text": f"6月，倒查的{discovery_total}个问题中，上报{discovery_reported}个，问题发现率{discovery_rate}；应处置{should}件案件，已{('全部按期处置' if should == done else f'按期处置{done}件')}，处置率{rate if rate else '100.00%'}{quality_text}。"
    }


def build_law_enforcement(street_name: str, tables: dict[str, list[TableSheet]]) -> dict[str, object]:
    doc_tables = _find_doc_tables(tables, "综合执法")
    if not doc_tables:
        return {"appeal_text": "", "key_assignment_text": ""}
    appeal_table = doc_tables[0] if len(doc_tables) > 0 else None
    key_table = doc_tables[4] if len(doc_tables) > 4 else None
    appeal_text = ""
    key_text = ""
    if appeal_table:
        rows = [row for row in appeal_table.rows[1:] if row and row[1] == street_name]
        if rows:
            row = rows[0]
            categories = [
                ("施工工地管理", row[5] if len(row) > 5 else 0),
                ("占道经营", row[6] if len(row) > 6 else 0),
                ("公共区域环境秩序", row[3] if len(row) > 3 else 0),
                ("非法小广告", row[2] if len(row) > 2 else 0),
                ("露天烧烤（含露天焚烧）", row[5 - 1] if len(row) > 4 else 0),
                ("生活垃圾分类", row[4] if len(row) > 4 else 0),
            ]
            # correct order with explicit columns
            categories = [
                ("非法小广告", row[2] if len(row) > 2 else 0),
                ("公共区域环境秩序", row[3] if len(row) > 3 else 0),
                ("生活垃圾分类", row[4] if len(row) > 4 else 0),
                ("露天烧烤（含露天焚烧）", row[5] if len(row) > 5 else 0),
                ("施工工地管理", row[6] if len(row) > 6 else 0),
                ("占道经营", row[7] if len(row) > 7 else 0),
            ]
            total = sum(int(float(v or 0)) for _, v in categories)
            parts = [f"{name}{fmt_int(v)}件" for name, v in sorted(categories, key=lambda x: int(float(x[1] or 0)), reverse=True) if int(float(v or 0)) > 0]
            appeal_text = f"6月，受理环境秩序问题诉求{total}件，分别为{'、'.join(parts)}。" if parts else f"6月，受理环境秩序问题诉求{total}件。"
    if key_table:
        rows = [row for row in key_table.rows[1:] if row and row[1] == street_name]
        if rows:
            row = rows[0]
            central = int(float(row[2] or 0))
            city = int(float(row[4] or 0))
            district = int(float(row[6] or 0))
            total = int(float(row[8] or 0))
            if total > 0:
                parts = []
                if central:
                    parts.append(f"中央{central}件")
                if city:
                    parts.append(f"市级{city}件")
                if district:
                    parts.append(f"区级{district}件")
                if parts == ["市级1件"]:
                    key_text = "有1件市级重点交办事项（发单交办），未发现未按要求完成重点交办事项情况。"
                else:
                    prefix = "、".join(parts) if parts else "中央、市、区级"
                    key_text = f"有{total}件{prefix}重点交办事项（发单交办），未发现未按要求完成重点交办事项情况。"
            else:
                key_text = "无中央、市、区级重点交办事项（发单交办）。"
    return {"appeal_text": appeal_text, "key_assignment_text": key_text}


def build_suggestions(*sections: dict[str, object]) -> list[dict[str, object]]:
    normal, waste, furniture, backstreet, grid, law = sections
    suggestions: list[dict[str, object]] = []
    if normal.get("problem_discovery", {}).get("current_month_total", 0):
        suggestions.append({"number": "一是", "title": "", "body": SUGGESTION_TEXTS["normal"]})
    if waste.get("city_check_text") or waste.get("district_check_text"):
        suggestions.append({"number": "二是", "title": "", "body": SUGGESTION_TEXTS["waste"]})
    if furniture.get("text") and len(suggestions) < 3:
        suggestions.append({"number": "三是", "title": "", "body": SUGGESTION_TEXTS["furniture"]})
    if grid.get("text") and len(suggestions) < 3:
        suggestions.append({"number": "三是", "title": "", "body": SUGGESTION_TEXTS["grid"]})
    if backstreet.get("text") and len(suggestions) < 3:
        suggestions.append({"number": "三是", "title": "", "body": SUGGESTION_TEXTS["backstreet"]})
    return suggestions[:3]


def generate_batch(
    input_paths: list[Path],
    template_dir: Path,
    output_dir: Path,
    progress: Callable[[int, int, str], None] | None = None,
    detail_template: Path | None = None,
    report_template: Path | None = None,
) -> list[Path]:
    sources = _iter_source_files(input_paths)
    wb_path = _find_source(sources, {".xlsx"}, ["现场检查用数据"])
    if wb_path is None:
        raise FileNotFoundError("未找到1.6月现场检查用数据.xlsx")

    street_summaries, data_rows, _ = _load_workbook_rows(wb_path)
    period = _detect_period(data_rows)
    doc_paths = [p for p in sources if p.suffix.lower() in {".docx", ".doc", ".wps"}]
    tables = _load_tables_by_name(doc_paths)
    report_tables = _collect_report_tables(doc_paths, progress)

    summaries = list(street_summaries.values())
    if not summaries:
        raise ValueError("无法从表1识别街镇列表")

    resolved_detail_template = detail_template or (template_dir / "街道案件明细表模板.xlsx")
    resolved_report_template = report_template or (template_dir / "街道环境建设管理工作运行情况分析报告模板.docx")
    if not resolved_detail_template.exists():
        raise FileNotFoundError(str(resolved_detail_template))
    if not resolved_report_template.exists():
        raise FileNotFoundError(str(resolved_report_template))

    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[Path] = []
    total = len(summaries)
    for idx, summary in enumerate(summaries, start=1):
        street = str(summary["street_name"])
        if progress:
            progress(idx - 1, total, f"正在生成 {street}")
        context = build_street_context(street, summary, data_rows, period, tables)
        street_dir = output_dir / street
        street_dir.mkdir(parents=True, exist_ok=True)
        detail_rows = [
            normalize_detail_row(r)
            for r in data_rows
            if str(r.get("月份") or "") == period[4] and str(r.get("检查点位2级") or "") == street
        ]
        detail_path = street_dir / f"{period[4]}{street}案件明细表.xlsx"
        report_path = street_dir / f"{period[4]}{street}环境建设管理工作运行情况分析报告.docx"
        render_detail_workbook(resolved_detail_template, detail_path, detail_rows)
        render_docx_template(resolved_report_template, report_path, asdict(context))
        results.extend([detail_path, report_path])
        if progress:
            progress(idx, total, f"已生成 {street}")

    summary_path = output_dir / "专项表格汇总.xlsx"
    render_summary_workbook(summary_path, report_tables)
    results.append(summary_path)
    return results


def generate_summary_only(
    input_paths: list[Path],
    output_dir: Path,
    progress: Callable[[int, int, str], None] | None = None,
) -> list[Path]:
    sources = _iter_source_files(input_paths)
    doc_paths = [p for p in sources if p.suffix.lower() in {".docx", ".doc", ".wps"}]
    if not doc_paths:
        raise ValueError("未找到可提取表格的 Word/WPS 文件")

    report_tables = _collect_report_tables(doc_paths, progress)
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "专项表格汇总.xlsx"
    render_summary_workbook(summary_path, report_tables)
    return [summary_path]


def _load_tables_by_name(doc_paths: list[Path]) -> dict[str, list[TableSheet]]:
    tables: dict[str, list[TableSheet]] = {}
    for doc in doc_paths:
        tables[doc.name] = extract_tables(doc)
    return tables


def _collect_report_tables(
    doc_paths: list[Path],
    progress: Callable[[int, int, str], None] | None = None,
) -> list[TableSheet]:
    report_tables: list[TableSheet] = []
    total = len(doc_paths)
    for idx, doc in enumerate(doc_paths, start=1):
        if progress:
            progress(idx - 1, total, f"正在提取表格 {doc.name}")
        report_tables.extend(extract_tables(doc))
        if progress:
            progress(idx, total, f"已提取完成 {doc.name}")
    return report_tables


def normalize_detail_row(row: dict[str, object]) -> dict[str, object]:
    out = {k: row.get(k, "") for k in DETAIL_COLUMN_KEYS}
    out["上报时间"] = fmt_datetime(out.get("上报时间"))
    out["截止时间"] = fmt_datetime(out.get("截止时间"))
    out["结案时间"] = fmt_datetime(out.get("结案时间"))
    out["检查点位1级"] = out.get("检查点位1级", "")
    out["检查点位2级"] = out.get("检查点位2级", "")
    out["检查点位3级"] = out.get("检查点位3级", "")
    out["检查指标1级"] = out.get("检查指标1级", "")
    out["检查指标2级"] = out.get("检查指标2级", "")
    out["检查指标3级"] = out.get("检查指标3级", "")
    return out
